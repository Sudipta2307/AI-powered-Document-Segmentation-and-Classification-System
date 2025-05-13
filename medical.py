import streamlit as st
import os
import re
import logging
import numpy as np
import pandas as pd
import fitz  # PyMuPDF
import umap
import hdbscan
import io
import base64
import json
import google.generativeai as genai
from typing import List, Dict, Any, Optional, Tuple
from sentence_transformers import SentenceTransformer
from sklearn.preprocessing import StandardScaler
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
import docx2txt
import openpyxl
import csv
from PyPDF2 import PdfReader


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


GOOGLE_API_KEY = "AIzaSyD4IQzP34KF_VI3rUYhuGumIAhYLBgAHbg"
genai.configure(api_key=GOOGLE_API_KEY)

class DocumentProcessor:
    """Base class for processing various document types"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from a document file"""
        raise NotImplementedError("Subclasses must implement extract_text method")

class PDFProcessor(DocumentProcessor):
    """Process PDF documents"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from a PDF file"""
        text = ""
        try:
            # First try with PyMuPDF for better formatting
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text()
        except Exception as e:
            logger.warning(f"PyMuPDF failed: {e}, falling back to PyPDF2")
            # Fallback to PyPDF2
            with open(file_path, 'rb') as f:
                pdf_reader = PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text()
        return text
    
    def get_page_texts(self, file_path: str) -> List[str]:
        """Extract text from each page of a PDF file"""
        try:
            doc = fitz.open(file_path)
            return [page.get_text() for page in doc]
        except Exception as e:
            logger.warning(f"Error extracting pages with PyMuPDF: {e}")
            
            page_texts = []
            with open(file_path, 'rb') as f:
                pdf_reader = PdfReader(f)
                for page in pdf_reader.pages:
                    page_texts.append(page.extract_text())
            return page_texts

class WordProcessor(DocumentProcessor):
    """Process Word documents"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from a Word document"""
        return docx2txt.process(file_path)

class ExcelProcessor(DocumentProcessor):
    """Process Excel documents"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from an Excel file"""
        text = ""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            text += f"Sheet: {sheet}\n"
            
            for row in worksheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                text += f"{row_text}\n"
            text += "\n"
            
        return text

class CSVProcessor(DocumentProcessor):
    """Process CSV documents"""
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from a CSV file"""
        text = ""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for row in reader:
                text += " | ".join(row) + "\n"
        return text

class DocumentFactory:
    """Factory for creating document processors"""
    
    @staticmethod
    def get_processor(file_extension: str) -> DocumentProcessor:
        """Get appropriate document processor based on file extension"""
        processors = {
            'pdf': PDFProcessor(),
            'docx': WordProcessor(),
            'doc': WordProcessor(),
            'xlsx': ExcelProcessor(),
            'xls': ExcelProcessor(),
            'csv': CSVProcessor()
        }
        
        processor = processors.get(file_extension.lower())
        if processor is None:
            raise ValueError(f"Unsupported file type: {file_extension}")
        
        return processor

class AdvancedMedicalRecordProcessor:
    def __init__(self):
        """
        Advanced Medical Record Processing System
        """
        # Load high-performance embedding model
        self.embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
        
        # Medical terminology extraction patterns
        self.medical_patterns = {
            'diagnosis': [
                r'\b(diagnosis|diagnosed with|condition|finding):\s*([^\n]+)',
                r'\b(primary|secondary|working) diagnosis:\s*([^\n]+)'
            ],
            'medications': [
                r'\b(medication|prescription|drug):\s*([^\n]+)',
                r'\b(current medications?|prescribed):\s*([^\n]+)'
            ],
            'procedures': [
                r'\b(procedure|surgery|intervention):\s*([^\n]+)',
                r'\b(performed|conducted):\s*([^\n]+)'
            ]
        }
    
    def advanced_text_analysis(self, text: str) -> Dict[str, List[str]]:
        """
        Advanced medical text analysis
        
        Args:
            text (str): Input medical text
        
        Returns:
            Dict of extracted medical information
        """
        analysis_results = {
            'diagnoses': [],
            'medications': [],
            'procedures': [],
            'key_entities': [],
            'sentiment': self._analyze_medical_sentiment(text)
        }
        
        # Extract medical information
        for category, patterns in self.medical_patterns.items():
            for pattern in patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    category_key = category
                    if category.endswith('s'):
                        category_key = category[:-1]
                    analysis_results[category_key] = [
                        match[1].strip() if isinstance(match, tuple) else match.strip() 
                        for match in matches
                    ]
        
        # Named Entity Recognition (Basic Implementation)
        analysis_results['key_entities'] = self._extract_key_entities(text)
        
        return analysis_results
    
    def _analyze_medical_sentiment(self, text: str) -> str:
        """
        Basic medical text sentiment analysis
        
        Args:
            text (str): Input medical text
        
        Returns:
            Sentiment classification
        """
        # Basic sentiment indicators for medical context
        negative_indicators = [
            'critical', 'severe', 'emergency', 'urgent', 
            'complications', 'deteriorating', 'risk'
        ]
        positive_indicators = [
            'improving', 'recovery', 'healing', 'stable', 
            'progress', 'positive', 'good condition'
        ]
        
        text_lower = text.lower()
        
        # Count sentiment indicators
        negative_count = sum(1 for indicator in negative_indicators if indicator in text_lower)
        positive_count = sum(1 for indicator in positive_indicators if indicator in text_lower)
        
        # Sentiment classification
        if negative_count > positive_count:
            return 'Concerning'
        elif positive_count > negative_count:
            return 'Optimistic'
        else:
            return 'Neutral'
    
    def _extract_key_entities(self, text: str) -> List[str]:
        """
        Extract key medical entities
        
        Args:
            text (str): Input text
        
        Returns:
            List of key entities
        """
        # Regex patterns for medical entities
        entity_patterns = [
            r'\b([A-Z][a-z]+ [A-Z][a-z]+)\b',  # Potential names
            r'\b(\d+\s*(?:mg|g|ml|cc))\b',     # Dosage
            r'\b([A-Z]{2,})\b',                # Abbreviations
        ]
        
        entities = []
        for pattern in entity_patterns:
            entities.extend(re.findall(pattern, text))
        
        return list(set(entities))
    
    def generate_medical_summary(self, clusters: Dict[str, Any]) -> str:
        """
        Generate comprehensive medical document summary
        
        Args:
            clusters (Dict): Clustering results
        
        Returns:
            Detailed medical document summary
        """
        summary = "Medical Document Analysis Report\n"
        summary += "=" * 40 + "\n\n"
        
        # Overview
        summary += f"Total Pages: {len(clusters['pages'])}\n"
        summary += f"Distinct Clusters: {clusters['num_clusters']}\n\n"
        
        # Detailed Cluster Analysis
        for cluster_id in set(clusters['clusters']):
            if cluster_id == -1:  # Skip noise cluster
                continue
                
            summary += f"Cluster {cluster_id} Analysis:\n"
            summary += "-" * 20 + "\n"
            
            # Collect pages in this cluster
            cluster_pages = [
                page for i, page in enumerate(clusters['pages']) 
                if clusters['clusters'][i] == cluster_id
            ]
            
            # Aggregate analysis
            cluster_analysis = self._aggregate_cluster_analysis(cluster_pages)
            
            summary += f"Pages in Cluster: {len(cluster_pages)}\n"
            summary += f"Predominant Sentiment: {cluster_analysis['sentiment']}\n"
            
            if cluster_analysis['diagnoses']:
                summary += "Key Diagnoses:\n"
                for diag in cluster_analysis['diagnoses']:
                    summary += f"  - {diag}\n"
            
            if cluster_analysis['medications']:
                summary += "Medications Mentioned:\n"
                for med in cluster_analysis['medications']:
                    summary += f"  - {med}\n"
            
            summary += "\n"
        
        return summary
    
    def _aggregate_cluster_analysis(self, pages: List[str]) -> Dict[str, Any]:
        """
        Aggregate analysis across cluster pages
        
        Args:
            pages (List[str]): Pages in a cluster
        
        Returns:
            Aggregated analysis
        """
        aggregated = {
            'diagnoses': [],
            'medications': [],
            'procedures': [],
            'sentiment': 'Neutral'
        }
        
        # Analyze each page
        for page in pages:
            page_analysis = self.advanced_text_analysis(page)
            
            # Aggregate findings
            aggregated['diagnoses'].extend(page_analysis.get('diagnoses', []))
            aggregated['medications'].extend(page_analysis.get('medications', []))
            aggregated['procedures'].extend(page_analysis.get('procedures', []))
        
        # Remove duplicates
        for key in ['diagnoses', 'medications', 'procedures']:
            aggregated[key] = list(set(aggregated[key]))
        
        return aggregated

class MedicalRecordClustering:
    def __init__(self):
        """
        Advanced Medical Record Clustering System
        """
        self.processor = AdvancedMedicalRecordProcessor()
        self.embedding_model = self.processor.embedding_model
    
    def cluster_medical_records(self, file_path: str) -> Dict[str, Any]:
        """
        Advanced clustering with enhanced features
        
        Args:
            file_path (str): Path to document file
        
        Returns:
            Clustering results with advanced analysis
        """
        # Get file extension
        file_extension = os.path.splitext(file_path)[1][1:]
        
        # Create appropriate processor
        doc_processor = DocumentFactory.get_processor(file_extension)
        
        # Extract text
        if file_extension == 'pdf':
            # Special handling for PDFs to get page-by-page text
            pdf_processor = doc_processor
            page_texts = pdf_processor.get_page_texts(file_path)
        else:
            # For other document types, treat as a single page
            full_text = doc_processor.extract_text(file_path)
            page_texts = [full_text]
        
        # Generate embeddings
        embeddings = self.embedding_model.encode(page_texts, convert_to_numpy=True)
        
        # Dimensionality Reduction (if more than 2 pages)
        if len(page_texts) > 2:
            reducer = umap.UMAP(
                n_components=min(5, len(page_texts) - 1),
                random_state=42, 
                metric='cosine'
            )
            reduced_embeddings = reducer.fit_transform(embeddings)
            
            # Clustering (if more than 2 pages)
            clusterer = hdbscan.HDBSCAN(
                min_cluster_size=2,
                min_samples=1,
                prediction_data=True
            )
            cluster_labels = clusterer.fit_predict(reduced_embeddings)
        else:
            # Not enough pages for meaningful clustering
            cluster_labels = np.zeros(len(page_texts), dtype=int)
        
        # Convert numpy types to Python native types for JSON compatibility
        cluster_labels_list = [int(label) for label in cluster_labels]
        
        # Prepare results
        results = {
            'clusters': cluster_labels_list,
            'num_clusters': len(set([c for c in cluster_labels_list if c != -1])),
            'pages': page_texts,
            'detailed_analysis': []
        }
        
        # Generate detailed analysis for each page
        for i, page_text in enumerate(page_texts):
            page_analysis = self.processor.advanced_text_analysis(page_text)
            page_analysis['cluster'] = int(cluster_labels[i])  # Convert to Python int
            results['detailed_analysis'].append(page_analysis)
        
        return results

class DocumentChatSystem:
    """Chat-based interface for document Q&A using Google's Gemini API"""
    
    def __init__(self, model_name: str = "all-MiniLM-L6-v2"):
        """Initialize chat system with embedding model"""
        self.embeddings = SentenceTransformer(model_name)
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000, 
            chunk_overlap=200
        )
        self.vector_store = None
        self.full_document_text = ""
        
        # Initialize Gemini model
        self.gemini_model = genai.GenerativeModel('gemini-1.5-flash')
    
    def process_document(self, file_path: str) -> None:
        """Process document for Q&A"""
        # Get file extension
        file_extension = os.path.splitext(file_path)[1][1:]
        
        # Create appropriate processor
        doc_processor = DocumentFactory.get_processor(file_extension)
        
        # Extract text
        document_text = doc_processor.extract_text(file_path)
        self.full_document_text = document_text
        
        # Split into chunks
        chunks = self.text_splitter.split_text(document_text)
        
        # Create vector store
        self.vector_store = FAISS.from_texts(
            chunks, 
            embedding=self.get_embedding_wrapper()
        )
        
        # Save for future use
        self.vector_store.save_local("faiss_index")
    
    def get_embedding_wrapper(self):
        """Create a wrapper for sentence_transformers model to work with LangChain"""
        class EmbeddingWrapper:
            def __init__(self, model):
                self.model = model
                
            def embed_documents(self, texts):
                return self.model.encode(texts).tolist()
                
            def embed_query(self, text):
                return self.model.encode([text])[0].tolist()
                
            # Make the wrapper callable for compatibility
            def __call__(self, texts):
                if isinstance(texts, str):
                    return self.embed_query(texts)
                else:
                    return self.embed_documents(texts)
        
        return EmbeddingWrapper(self.embeddings)
    
    def answer_question(self, question: str) -> str:
        """Answer question based on document content using Gemini"""
        if self.vector_store is None:
            if os.path.exists("faiss_index"):
                self.vector_store = FAISS.load_local(
                    "faiss_index", 
                    self.get_embedding_wrapper()
                )
            else:
                return "Please process a document first"
        
        try:
            # Retrieve relevant chunks
            docs = self.vector_store.similarity_search(question, k=5)
            
            # Create context string
            context = "\n\n".join([doc.page_content for doc in docs])
            
            # Prepare prompt for Gemini
            prompt = f"""
            You are an expert medical document analyst. Answer the following question based only on the provided context from a medical document.
            
            CONTEXT:
            {context}
            
            QUESTION:
            {question}
            
            Provide a clear, concise, and accurate answer based solely on the information in the context. 
            If the context doesn't contain relevant information to answer the question, state that you cannot 
            answer based on the available information.
            """
            
            # Get response from Gemini
            response = self.gemini_model.generate_content(prompt)
            answer = response.text
            
            return answer
            
        except Exception as e:
            logger.error(f"Error using Gemini API: {str(e)}", exc_info=True)
            return f"I encountered an error while processing your question. Please try again or rephrase your question. Error details: {str(e)}"

def main():
    st.set_page_config(
        page_title="Medical Document Analyzer & Chat",
        page_icon="🩺",
        layout="wide"
    )
    
    st.title("🩺 Advanced Medical Document Intelligence System")
    
    # Initialize session state
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    if 'current_document' not in st.session_state:
        st.session_state.current_document = None
    if 'chat_system' not in st.session_state:
        st.session_state.chat_system = DocumentChatSystem()
    
    # Sidebar for document upload and options
    with st.sidebar:
        st.header("📄 Document Upload")
        
        uploaded_file = st.file_uploader(
            "Upload Medical Document", 
            type=['pdf', 'docx', 'doc', 'xlsx', 'xls', 'csv'],
            help="Upload a medical document for analysis"
        )
        
        st.header("⚙️ Analysis Options")
        
        analysis_type = st.radio(
            "Select Analysis Type",
            options=["Document Analysis", "Chat with Document"],
            help="Choose how to analyze your document"
        )
        
        if uploaded_file:
            # Create temp directory if it doesn't exist
            os.makedirs('temp', exist_ok=True)
            
            # Save the uploaded file
            temp_path = os.path.join('temp', uploaded_file.name)
            with open(temp_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())
            
            # Process button
            if st.button("Process Document"):
                with st.spinner("Processing document..."):
                    try:
                        # Process document for both analysis types
                        if analysis_type == "Document Analysis":
                            # Process for clustering and analysis
                            clustering_model = MedicalRecordClustering()
                            st.session_state.results = clustering_model.cluster_medical_records(temp_path)
                            st.session_state.current_document = temp_path
                            st.success(f"Document processed successfully!")
                        
                        # Always process for chat capability
                        st.session_state.chat_system.process_document(temp_path)
                        st.session_state.current_document = temp_path
                        
                    except Exception as e:
                        st.error(f"Error processing document: {str(e)}")
                        logger.error(f"Document processing error: {str(e)}", exc_info=True)
        
        # Add Gemini API configuration section
        with st.expander("API Configuration"):
            api_key = st.text_input("Google API Key", value=GOOGLE_API_KEY, type="password")
            if st.button("Update API Key"):
                genai.configure(api_key=api_key)
                st.success("API key updated successfully!")
    
    # Main content area
    if analysis_type == "Document Analysis":
        if st.session_state.get('results'):
            display_document_analysis()
    else:  # Chat with Document
        display_chat_interface()
    
    # Cleanup temp files on session end
    for file in os.listdir('temp'):
        try:
            os.remove(os.path.join('temp', file))
        except Exception as e:
            logger.warning(f"Failed to clean up {file}: {str(e)}")

def display_document_analysis():
    """Display document analysis results"""
    results = st.session_state.results
    
    # Generate summary
    processor = AdvancedMedicalRecordProcessor()
    summary = processor.generate_medical_summary(results)
    
    # Visualization and Insights
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("📊 Document Clustering")
        # Visualize clusters
        cluster_counts = {}
        for cluster in results['clusters']:
            if cluster != -1:  # Skip noise points
                cluster_counts[f"Cluster {cluster}"] = cluster_counts.get(f"Cluster {cluster}", 0) + 1
        
        # Add noise points if any
        noise_count = results['clusters'].count(-1)
        if noise_count > 0:
            cluster_counts["Unclustered"] = noise_count
        
        if cluster_counts:
            st.bar_chart(cluster_counts)
        else:
            st.info("No clusters found. Document may be too short for meaningful clustering.")
    
    with col2:
        st.subheader("📝 Quick Insights")
        st.metric("Total Pages/Sections", len(results['pages']))
        st.metric("Distinct Clusters", results['num_clusters'])
        
        # Detect document type
        file_extension = os.path.splitext(st.session_state.current_document)[1][1:]
        st.metric("Document Type", file_extension.upper())
    
    # Detailed Analysis Tabs
    tab1, tab2, tab3 = st.tabs([
        "Summary", 
        "Detailed Analysis", 
        "Download Options"
    ])
    
    with tab1:
        st.markdown(summary)
    
    with tab2:
        for i, analysis in enumerate(results['detailed_analysis'], 1):
            with st.expander(f"Page/Section {i} Analysis"):
                # Show cluster assignment
                st.write(f"**Cluster:** {'Unclustered' if analysis['cluster'] == -1 else analysis['cluster']}")
                
                # Show sentiment
                sentiment = analysis['sentiment']
                sentiment_color = {
                    'Concerning': 'red',
                    'Neutral': 'blue',
                    'Optimistic': 'green'
                }.get(sentiment, 'black')
                st.write(f"**Sentiment:** :{sentiment_color}[{sentiment}]")
                
                # Show key findings
                if analysis.get('diagnoses'):
                    st.write("**Diagnoses:**")
                    for diag in analysis['diagnoses']:
                        st.write(f"- {diag}")
                
                if analysis.get('medications'):
                    st.write("**Medications:**")
                    for med in analysis['medications']:
                        st.write(f"- {med}")
                
                if analysis.get('procedures'):
                    st.write("**Procedures:**")
                    for proc in analysis['procedures']:
                        st.write(f"- {proc}")
                
                if analysis.get('key_entities'):
                    st.write("**Key Entities:**")
                    for entity in analysis['key_entities'][:10]:  # Limit to top 10
                        st.write(f"- {entity}")
    
    with tab3:
        # Download options
        summary_file = io.BytesIO(summary.encode())
        st.download_button(
            label="📄 Download Summary Report",
            data=summary_file,
            file_name='medical_document_summary.md',
            mime='text/markdown'
        )
        
        # JSON export - With JSON serialization fix
        # Create a serializable copy of results
        serializable_results = {
            'clusters': results['clusters'],
            'num_clusters': results['num_clusters'],
            'pages': results['pages'],
            'detailed_analysis': []
        }
        
        # Convert any NumPy types to native Python types
        for analysis in results['detailed_analysis']:
            serializable_analysis = {}
            for key, value in analysis.items():
                # Convert NumPy types to Python native types
                if isinstance(value, np.integer):
                    serializable_analysis[key] = int(value)
                elif isinstance(value, np.floating):
                    serializable_analysis[key] = float(value)
                elif isinstance(value, np.ndarray):
                    serializable_analysis[key] = value.tolist()
                else:
                    serializable_analysis[key] = value
            serializable_results['detailed_analysis'].append(serializable_analysis)
        
        json_results = json.dumps(serializable_results, indent=2)
        json_file = io.BytesIO(json_results.encode())
        st.download_button(
            label="📊 Download Full Analysis",
            data=json_file,
            file_name='medical_document_analysis.json',
            mime='application/json'
        )

def display_chat_interface():
    """Display chat interface for document Q&A"""
    st.subheader("💬 Chat with Your Document ")
    
    # Check if document is processed
    if not st.session_state.current_document:
        st.info("Please upload and process a document first")
        return
    
    # Display chat history
    for i, (question, answer) in enumerate(st.session_state.chat_history):
        with st.chat_message("user"):
            st.write(question)
        with st.chat_message("assistant", avatar="🩺"):
            st.write(answer)
    
    # Chat input
    user_question = st.chat_input("Ask a question about your medical document...")
    
    if user_question:
        # Add user question to chat
        with st.chat_message("user"):
            st.write(user_question)
        
        # Get answer from Gemini
        with st.spinner("Analyzing document and generating response..."):
            try:
                answer = st.session_state.chat_system.answer_question(user_question)
                
                # Display answer
                with st.chat_message("assistant", avatar="🩺"):
                    st.write(answer)
                
                # Add to chat history
                st.session_state.chat_history.append((user_question, answer))
                
            except Exception as e:
                st.error(f"Error generating answer: {str(e)}")
                logger.error(f"Chat error: {str(e)}", exc_info=True)

if __name__ == '__main__':
    # Ensure temp directory exists
    os.makedirs('temp', exist_ok=True)
    main()